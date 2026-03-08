import os
import logging
from pathlib import Path
from typing import Dict, Any, List
import pypdf
from pypdf import PdfWriter, PdfReader
from pypdf.generic import NameObject, TextStringObject
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import json

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class FormFiller:
    """Fill form templates (PDF and Excel) with extracted data"""
    
    def __init__(self):
        self.output_dir = Path("../outputs")
        self.output_dir.mkdir(exist_ok=True)
    
    def fill_template(self, template_path: str, data: Dict[str, Any]) -> str:
        """Fill a form template with data and return output path"""
        template_path = Path(template_path)
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        try:
            if template_path.suffix.lower() == '.pdf':
                return self._fill_pdf_template(template_path, data)
            elif template_path.suffix.lower() in ['.xlsx', '.xls']:
                return self._fill_excel_template(template_path, data)
            else:
                raise ValueError(f"Unsupported template format: {template_path.suffix}")
                
        except Exception as e:
            logger.error(f"Failed to fill template {template_path}: {e}")
            raise
    
    def _fill_pdf_template(self, template_path: Path, data: Dict[str, Any]) -> str:
        """Fill PDF form fields with data"""
        try:
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"filled_{template_path.stem}_{timestamp}.pdf"
            output_path = self.output_dir / output_filename
            
            # Read the template
            reader = PdfReader(template_path)
            writer = PdfWriter()
            
            # Copy all pages
            for page in reader.pages:
                writer.add_page(page)
            
            # Fill form fields if they exist
            if '/AcroForm' in reader.trailer['/Root']:
                fields = reader.get_fields()
                
                if fields:
                    logger.info(f"Found {len(fields)} form fields in PDF")
                    
                    # Update form fields
                    for field_name, field_value in data.items():
                        # Try to find matching field
                        matching_field = self._find_matching_pdf_field(field_name, list(fields.keys()))
                        
                        if matching_field:
                            writer.update_page_form_field_values(
                                writer.pages[0],  # Assuming single form for now
                                {matching_field: str(field_value)}
                            )
                            logger.info(f"Filled PDF field '{matching_field}' with '{field_value}'")
            
            # Write the output
            with open(output_path, 'wb') as output_file:
                writer.write(output_file)
            
            logger.info(f"PDF form saved to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"PDF filling failed: {e}")
            # Fallback: create a new PDF with data overlay
            return self._create_pdf_with_data_overlay(template_path, data)
    
    def _fill_excel_template(self, template_path: Path, data: Dict[str, Any]) -> str:
        """Fill Excel template with data"""
        try:
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"filled_{template_path.stem}_{timestamp}.xlsx"
            output_path = self.output_dir / output_filename
            
            # Load the template
            workbook = openpyxl.load_workbook(template_path)
            
            # Process each sheet
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                self._fill_excel_sheet(sheet, data)
            
            # Save the filled workbook
            workbook.save(output_path)
            workbook.close()
            
            logger.info(f"Excel form saved to {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Excel filling failed: {e}")
            raise
    
    def _fill_excel_sheet(self, sheet, data: Dict[str, Any]):
        """Fill data into an Excel sheet"""
        # First, try to map data to column headers
        headers = {}
        header_row = None
        
        # Find header row (usually first row with data)
        for row_idx in range(1, min(10, sheet.max_row + 1)):  # Check first 10 rows
            row_has_headers = False
            for col_idx, cell in enumerate(sheet[row_idx], 1):
                if cell.value and str(cell.value).strip():
                    header_name = str(cell.value).strip()
                    if self._is_field_like(header_name):
                        headers[header_name.lower()] = col_idx
                        row_has_headers = True
            
            if row_has_headers:
                header_row = row_idx
                break
        
        if not headers:
            logger.warning("No headers found in Excel sheet")
            return
        
        logger.info(f"Found {len(headers)} headers in Excel sheet")
        
        # Fill data in the first empty row
        data_row = header_row + 1
        for field_name, field_value in data.items():
            # Find matching column
            matching_column = self._find_matching_excel_column(field_name, list(headers.keys()))
            
            if matching_column:
                col_idx = headers[matching_column]
                sheet.cell(row=data_row, column=col_idx, value=field_value)
                logger.info(f"Filled Excel cell ({data_row}, {col_idx}) with '{field_value}'")
    
    def _find_matching_pdf_field(self, data_field: str, pdf_fields: List[str]) -> str:
        """Find matching PDF field name for data field"""
        data_field_lower = data_field.lower()
        
        # Exact match
        for pdf_field in pdf_fields:
            if pdf_field.lower() == data_field_lower:
                return pdf_field
        
        # Partial match
        for pdf_field in pdf_fields:
            if data_field_lower in pdf_field.lower() or pdf_field.lower() in data_field_lower:
                return pdf_field
        
        # Synonym matching
        synonyms = {
            "name": ["full_name", "applicant", "person"],
            "email": ["email_address", "mail"],
            "phone": ["phone_number", "telephone", "mobile"],
            "address": ["street_address", "residence"],
            "date_of_birth": ["dob", "birth_date"],
            "pan": ["pan_number", "pan_no"],
            "aadhar": ["aadhaar", "uid"]
        }
        
        for canonical, synonym_list in synonyms.items():
            if data_field_lower in [canonical] + synonym_list:
                for pdf_field in pdf_fields:
                    pdf_field_lower = pdf_field.lower()
                    if pdf_field_lower in [canonical] + synonym_list:
                        return pdf_field
        
        return None
    
    def _find_matching_excel_column(self, data_field: str, headers: List[str]) -> str:
        """Find matching Excel column header for data field"""
        data_field_lower = data_field.lower()
        
        # Exact match
        for header in headers:
            if header == data_field_lower:
                return header
        
        # Partial match
        for header in headers:
            if data_field_lower in header or header in data_field_lower:
                return header
        
        # Synonym matching (same logic as PDF)
        synonyms = {
            "name": ["full_name", "applicant", "person"],
            "email": ["email_address", "mail"],
            "phone": ["phone_number", "telephone", "mobile"],
            "address": ["street_address", "residence"],
            "date_of_birth": ["dob", "birth_date"],
            "pan": ["pan_number", "pan_no"],
            "aadhar": ["aadhaar", "uid"]
        }
        
        for canonical, synonym_list in synonyms.items():
            if data_field_lower in [canonical] + synonym_list:
                for header in headers:
                    if header in [canonical] + synonym_list:
                        return header
        
        return None
    
    def _is_field_like(self, text: str) -> bool:
        """Check if text looks like a field name"""
        field_indicators = [
            'name', 'address', 'phone', 'email', 'date', 'signature',
            'number', 'id', 'code', 'amount', 'total', 'balance',
            'account', 'policy', 'invoice', 'receipt', 'reference'
        ]
        
        text_lower = text.lower()
        
        for indicator in field_indicators:
            if indicator in text_lower:
                return True
        
        return False
    
    def _create_pdf_with_data_overlay(self, template_path: Path, data: Dict[str, Any]) -> str:
        """Create a new PDF with data as text overlay (fallback method)"""
        try:
            from reportlab.pdfgen import canvas
            from reportlab.lib.pagesizes import letter
            
            # Create output filename
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_filename = f"overlay_{template_path.stem}_{timestamp}.pdf"
            output_path = self.output_dir / output_filename
            
            # Create a simple PDF with the data
            c = canvas.Canvas(str(output_path), pagesize=letter)
            
            # Add title
            c.setFont("Helvetica-Bold", 16)
            c.drawString(50, 750, "Extracted Data")
            
            # Add data
            c.setFont("Helvetica", 12)
            y_position = 700
            
            for field_name, field_value in data.items():
                if y_position < 50:  # Start new page if needed
                    c.showPage()
                    y_position = 700
                
                text = f"{field_name}: {field_value}"
                c.drawString(50, y_position, text)
                y_position -= 20
            
            c.save()
            
            logger.info(f"PDF overlay created at {output_path}")
            return str(output_path)
            
        except ImportError:
            logger.error("ReportLab not available for PDF overlay")
            # Fallback to JSON file
            return self._create_json_output(data, template_path.stem)
    
    def _create_json_output(self, data: Dict[str, Any], template_name: str) -> str:
        """Create JSON output as final fallback"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"data_{template_name}_{timestamp}.json"
        output_path = self.output_dir / output_filename
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"JSON output created at {output_path}")
        return str(output_path)
    
    def get_template_info(self, template_path: str) -> Dict[str, Any]:
        """Get information about a template"""
        template_path = Path(template_path)
        
        if not template_path.exists():
            raise FileNotFoundError(f"Template not found: {template_path}")
        
        info = {
            "name": template_path.name,
            "type": template_path.suffix.lower(),
            "size": template_path.stat().st_size,
            "fields": []
        }
        
        try:
            if template_path.suffix.lower() == '.pdf':
                reader = PdfReader(template_path)
                if '/AcroForm' in reader.trailer['/Root']:
                    fields = reader.get_fields()
                    if fields:
                        info["fields"] = list(fields.keys())
                        info["has_form_fields"] = True
                    else:
                        info["has_form_fields"] = False
                else:
                    info["has_form_fields"] = False
                    
            elif template_path.suffix.lower() in ['.xlsx', '.xls']:
                workbook = openpyxl.load_workbook(template_path, read_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    # Get headers from first row
                    for cell in sheet[1]:
                        if cell.value and str(cell.value).strip():
                            info["fields"].append(str(cell.value).strip())
                workbook.close()
                info["has_form_fields"] = len(info["fields"]) > 0
            
        except Exception as e:
            logger.error(f"Failed to analyze template {template_path}: {e}")
        
        return info
